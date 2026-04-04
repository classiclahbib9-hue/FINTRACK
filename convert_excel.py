"""
FinTrack – Excel Converter for Profit (1).xlsx
Reads the yearly sheets (2023-2026) and LAPTOP sheet,
converts everything to FinTrack transaction format,
and outputs a JSON file to inject into localStorage.
"""

import openpyxl
import json
import re
import random
import string
from datetime import datetime, date

EXCEL_PATH = r'C:\Users\PC\Desktop\Profit (2).xlsx'
OUTPUT_JSON = r'C:\Users\PC\software\1\import_data.json'

MONTH_MAP = {
    'january': 1, 'janvier': 1, 'jan': 1,
    'february': 2, 'fevrier': 2, 'feb': 2,
    'march': 3, 'mars': 3, 'mar': 3,
    'april': 4, 'avril': 4, 'apr': 4,
    'may': 5, 'mai': 5,
    'june': 6, 'juin': 6, 'jun': 6,
    'july': 7, 'juillet': 7, 'jul': 7,
    'august': 8, 'aout': 8, 'aug': 8,
    'september': 9, 'septembre': 9, 'sep': 9,
    'october': 10, 'octobre': 10, 'oct': 10,
    'november': 11, 'novembre': 11, 'nov': 11,
    'december': 12, 'decembre': 12, 'dec': 12,
}

# Income source label → FinTrack category
INCOME_CAT_MAP = {
    'logicielle': 'Freelance',
    'service': 'Freelance',
    'vente lic': 'Freelance',
    'vente laptop': 'Freelance',
    'montage': 'Freelance',
    'abonnement': 'Freelance',
    'side hustle': 'Freelance',
    'chantier': 'Freelance',
    'extra': 'Bonus',
    'mission': 'Bonus',
    'prime': 'Bonus',
    'material': 'Freelance',
    'income': 'Salary',
    'salary': 'Salary',
    'investment': 'Investment',
    'invest': 'Investment',
    'total profit': 'Freelance',
    'general profit': 'Freelance',
}

# Expense label → FinTrack category
EXPENSE_CAT_MAP = {
    'food': 'Food',
    'fast food': 'Food',
    'gym food': 'Food',
    'clothes': 'Shopping',
    'moto': 'Transport',
    'telecome': 'Health',
    'telecom': 'Health',
    'home': 'Housing',
    'health': 'Health',
    'electricite': 'Housing',
    'électricité': 'Housing',
    'tech': 'Shopping',
    'gaming': 'Entertainment',
    'other': 'Other',
    'investment': 'Investment',
    'invest': 'Investment',
    'other/books': 'Education',
    'annem': 'Other',
    'mahdi': 'Other',
    'baba': 'Other',
    'sada9a': 'Other',
    'wifey': 'Other',
    'sex': 'Entertainment',
}

transactions = []


def uid():
    return ''.join(random.choices(string.ascii_lowercase + string.digits, k=10))


def make_date(year, month, day=1):
    try:
        return date(int(year), int(month), min(int(day), 28)).isoformat()
    except Exception:
        return f"{int(year)}-{int(month):02d}-01"


def map_income_cat(label):
    l = str(label).lower().strip()
    for k, v in INCOME_CAT_MAP.items():
        if k in l:
            return v
    return 'Freelance'


def map_expense_cat(label):
    l = str(label).lower().strip()
    for k, v in EXPENSE_CAT_MAP.items():
        if k in l:
            return v
    return 'Other'


def safe_float(v):
    try:
        f = float(str(v).replace(' ', '').replace(',', '.'))
        if f > 0:
            return round(f, 2)
    except Exception:
        pass
    return None


def parse_yearly_sheet(ws, year):
    """
    Parse a yearly sheet (e.g. 2023, 2024, 2025, 2026).
    Structure: monthly blocks with 'total profit' income row and 'Liabilities' expense section.
    """
    rows = list(ws.iter_rows(values_only=True))
    current_month = None
    in_liabilities = False
    count = 0

    for i, row in enumerate(rows):
        # Detect month header
        for cell in row:
            if cell and isinstance(cell, str):
                cl = cell.lower().strip()
                for mname, mnum in MONTH_MAP.items():
                    if cl == mname or cl.startswith(mname):
                        current_month = mnum
                        in_liabilities = False
                        break

        if current_month is None:
            continue

        # Detect start of liabilities section
        for cell in row:
            if cell and isinstance(cell, str) and 'liabilit' in cell.lower():
                in_liabilities = True

        # --- INCOME: look for "total profit" row ---
        label = row[3] if len(row) > 3 else None   # col D usually has income labels
        value_col = row[4] if len(row) > 4 else None  # col E has amounts
        
        if label and isinstance(label, str):
            ll = label.lower().strip()
            if 'total profit' in ll or 'net' == ll:
                amount = safe_float(value_col)
                if amount and amount > 100:  # filter noise
                    transactions.append({
                        'id': uid(),
                        'type': 'income',
                        'amount': amount,
                        'category': 'Freelance',
                        'date': make_date(year, current_month),
                        'note': f'Total profit – {list(MONTH_MAP.keys())[list(MONTH_MAP.values()).index(current_month)].capitalize()} {year}'
                    })
                    count += 1

        # --- EXPENSES: from Liabilities section ---
        if in_liabilities:
            exp_label = row[6] if len(row) > 6 else None  # col G = expense category
            exp_val   = row[7] if len(row) > 7 else None  # col H = expense amount
            if exp_label and isinstance(exp_label, str):
                ll = exp_label.lower().strip()
                skip = ['total', 'liabilit', 'savings', 'profit', 'none', 'general',
                        'monthly', 'average', 'income', 'yearly', 'chiffre', 'prime']
                if not any(s in ll for s in skip) and ll not in ('', '-'):
                    amount = safe_float(exp_val)
                    if amount and 50 < amount < 5_000_000:
                        cat = map_expense_cat(ll)
                        transactions.append({
                            'id': uid(),
                            'type': 'expense',
                            'amount': amount,
                            'category': cat,
                            'date': make_date(year, current_month),
                            'note': f'{exp_label.strip()} – {year}'
                        })
                        count += 1

    print(f"  >> {year}: extracted {count} transactions")


def parse_laptop_sheet(ws):
    """
    LAPTOP sheet: each row is a sale with ARTICLE, BUY, SELL, MARGIN VALUE, DATE, CLIENT.
    We record MARGE VALEUR as Freelance income (the profit), and BUY as Shopping expense.
    """
    rows = list(ws.iter_rows(values_only=True))
    count = 0
    for i, row in enumerate(rows[1:], 1):  # skip header
        article  = row[0] if len(row) > 0 else None
        buy      = row[2] if len(row) > 2 else None
        sell     = row[3] if len(row) > 3 else None
        margin   = row[5] if len(row) > 5 else None
        raw_date = row[6] if len(row) > 6 else None
        client   = row[7] if len(row) > 7 else None

        if not article:
            continue

        # Parse date
        tx_date = None
        if isinstance(raw_date, datetime):
            tx_date = raw_date.date().isoformat()
        elif raw_date and isinstance(raw_date, str):
            m = re.match(r'(\d{1,2})[\/\-\.](\d{1,2})[\/\-\.](\d{2,4})', str(raw_date))
            if m:
                d, mo, yr = m.group(1), m.group(2), m.group(3)
                yr = '20' + yr if len(yr) == 2 else yr
                tx_date = f"{yr}-{mo.zfill(2)}-{d.zfill(2)}"
        if not tx_date:
            tx_date = '2025-01-01'  # fallback

        # Record profit (margin value) as income
        profit = safe_float(margin)
        if profit and profit > 0:
            note = str(article).strip()
            if client and isinstance(client, str):
                note += f' → {client.strip()}'
            transactions.append({
                'id': uid(),
                'type': 'income',
                'amount': profit,
                'category': 'Freelance',
                'date': tx_date,
                'note': note
            })
            count += 1

        # Record purchase cost as expense (only if sold)
        cost = safe_float(buy)
        sell_v = safe_float(sell)
        if cost and sell_v and cost > 0:
            transactions.append({
                'id': uid(),
                'type': 'expense',
                'amount': cost,
                'category': 'Shopping',
                'date': tx_date,
                'note': f'Purchase: {str(article).strip()}'
            })
            count += 1

    print(f"  >> LAPTOP: extracted {count} transactions")


# ── Main ─────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
print(f"Sheets found: {wb.sheetnames}")

for year in ['2023', '2024', '2025', '2026']:
    if year in wb.sheetnames:
        print(f"Processing {year}...")
        parse_yearly_sheet(wb[year], year)

if 'LAPTOP' in wb.sheetnames:
    print("Processing LAPTOP...")
    parse_laptop_sheet(wb['LAPTOP'])

# Sort by date newest-first (as expected by app)
transactions.sort(key=lambda t: t['date'], reverse=True)

print(f"\nOK Total transactions extracted: {len(transactions)}")
print("Sample (first 5) generated successfully")

# Save to CSV
import csv
with open('C:\\Users\\PC\\Desktop\\FinTrack_Backup.csv', 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)
    writer.writerow(['Date', 'Type', 'Category', 'Amount', 'Note']) # Headers
    for t in transactions:
        writer.writerow([t['date'], t['type'], t['category'], t['amount'], t['note']])

print(f"\nSaved to: C:\\Users\\PC\\Desktop\\FinTrack_Backup.csv")
